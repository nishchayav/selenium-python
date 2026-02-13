"""	Question:
	You have an Excel file salesdata.xlsx with a sheet named "2025" that contains:
	Product	Quantity	Price
	A	    10	        50
	B	    20	        30
	C	    15	        40
	Use Pandas to read the Excel sheet into a DataFrame.
	Add a new column "Total" which is Quantity * Price.
	Save the updated DataFrame to a new Excel file called salessummary.xlsx.
	(Bonus) Do the same using OpenPyXL without using Pandas		"""

import pandas as pd

# Read  file
df = pd.read_excel("salesdata.xlsx", sheet_name="2025")
# Add new column "Total"
df["Total"] = df["Quantity"] * df["Price"]
# Save to new Excel file
df.to_excel("salessummaryPANDAS.xlsx", index=False)

print("Sales summary file created successfully!(PANDAS)")



from openpyxl import load_workbook, Workbook
wb = load_workbook("salesdata.xlsx")
sheet= wb["2025"]

new_wb = Workbook()
new_sheet = new_wb.active
new_sheet.title = "Summary"
headers = ["Product", "Quantity", "Price", "Total"]
new_sheet.append(headers)

for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
    product, quantity, price = row
    total = quantity * price
    new_sheet.append([product, quantity, price, total])

new_wb.save("salessummaryOPENPYXL.xlsx")
print("Sales summary file created successfully!(OPENPYXL)")